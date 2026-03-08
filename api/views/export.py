import io
from datetime import datetime
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse


def _build_pdf(state, result, display_currency):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm,
                            leftMargin=2*cm, rightMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    # Title
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, spaceAfter=6)
    story.append(Paragraph('COZM Mobility Cost Estimate', title_style))

    scenario = state.get('scenarioName') or 'Unnamed Scenario'
    story.append(Paragraph(f'Scenario: {scenario}', styles['Normal']))
    story.append(Paragraph(f'Generated: {datetime.utcnow().strftime("%d %b %Y")}', styles['Normal']))
    story.append(Spacer(1, 0.5*cm))

    # Summary table
    total = result.get('totalAssignmentCost', 0)
    annual = result.get('annualAvgCost', 0)
    summary_data = [
        ['Total Assignment Cost', f'{display_currency} {total:,.0f}'],
        ['Annual Average Cost', f'{display_currency} {annual:,.0f}'],
        ['Home Country', state.get('homeCountryCode', '')],
        ['Host Country', state.get('hostCountryCode', '')],
        ['Tier', state.get('tier', '')],
        ['Duration', f"{state.get('startDate', '')} — {state.get('endDate', '')}"],
    ]
    summary_table = Table(summary_data, colWidths=[8*cm, 8*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f4f8')),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.5*cm))

    # Breakdown table
    story.append(Paragraph('Cost Breakdown', styles['Heading2']))
    breakdown = result.get('breakdown', [])
    if breakdown:
        headers = ['Category', 'Annual Amount', 'Total Amount', '%']
        rows = [headers]
        for item in breakdown:
            rows.append([
                item.get('category', ''),
                f"{display_currency} {item.get('amount', 0):,.0f}",
                f"{display_currency} {item.get('totalAmount', 0):,.0f}",
                f"{item.get('percentage', '0')}%",
            ])
        tbl = Table(rows, colWidths=[6*cm, 4*cm, 4*cm, 3*cm])
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('PADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(tbl)

    doc.build(story)
    buffer.seek(0)
    return buffer.read()


def _build_excel(state, result, display_currency):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = 'Cost Estimate'

    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    alt_fill = PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid')

    # Summary
    ws['A1'] = 'COZM Mobility Cost Estimate'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Scenario: {state.get('scenarioName', '')}"
    ws['A3'] = f"Generated: {datetime.utcnow().strftime('%d %b %Y')}"

    ws['A5'] = 'Summary'
    ws['A5'].font = Font(bold=True)
    summary = [
        ('Total Assignment Cost', f'{display_currency} {result.get("totalAssignmentCost", 0):,.0f}'),
        ('Annual Average Cost', f'{display_currency} {result.get("annualAvgCost", 0):,.0f}'),
        ('Home Country', state.get('homeCountryCode', '')),
        ('Host Country', state.get('hostCountryCode', '')),
        ('Tier', state.get('tier', '')),
    ]
    for i, (key, val) in enumerate(summary, start=6):
        ws.cell(row=i, column=1, value=key).font = Font(bold=True)
        ws.cell(row=i, column=2, value=val)

    # Breakdown headers
    start_row = 13
    ws.cell(row=start_row, column=1, value='Category').font = header_font
    ws.cell(row=start_row, column=1).fill = header_fill
    for col, h in enumerate(['Category', 'Annual Amount', 'Total Amount', '%'], start=1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for i, item in enumerate(result.get('breakdown', []), start=start_row + 1):
        fill = alt_fill if i % 2 == 0 else PatternFill()
        ws.cell(row=i, column=1, value=item.get('category', '')).fill = fill
        ws.cell(row=i, column=2, value=item.get('amount', 0)).fill = fill
        ws.cell(row=i, column=2).number_format = '#,##0'
        ws.cell(row=i, column=3, value=item.get('totalAmount', 0)).fill = fill
        ws.cell(row=i, column=3).number_format = '#,##0'
        ws.cell(row=i, column=4, value=f"{item.get('percentage', '0')}%").fill = fill

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 10

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_pdf(request):
    state = request.data.get('state', {})
    result = request.data.get('result', {})
    display_currency = request.data.get('displayCurrency', 'USD')

    try:
        pdf_bytes = _build_pdf(state, result, display_currency)
    except Exception as e:
        return Response({'error': 'PDF generation failed', 'details': str(e)},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    date_str = datetime.utcnow().strftime('%Y-%m-%d')
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="mobility-estimate-{date_str}.pdf"'
    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_excel(request):
    state = request.data.get('state', {})
    result = request.data.get('result', {})
    display_currency = request.data.get('displayCurrency', 'USD')

    try:
        xlsx_bytes = _build_excel(state, result, display_currency)
    except Exception as e:
        return Response({'error': 'Excel generation failed', 'details': str(e)},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    date_str = datetime.utcnow().strftime('%Y-%m-%d')
    response = HttpResponse(
        xlsx_bytes,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="mobility-estimate-{date_str}.xlsx"'
    return response
